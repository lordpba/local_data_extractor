import os
import json
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
from flask_cors import CORS
import hashlib
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io
import requests

# Load environment variables
load_dotenv()

from processor import (
    process_document,
    extract_structured_data_with_ollama
)
from models_config import (
    VISION_MODELS,
    HARDWARE_RECOMMENDATIONS,
    get_models_by_category,
    get_recommended_models_by_vram,
    get_all_model_names,
    get_family_tiers
)

app = Flask(__name__)
CORS(app)

# Configuration
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
UPLOAD_FOLDER = 'temp_uploads'
# CACHE_FOLDER = 'cache'  # Cache system disabled

# Ensure folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(CACHE_FOLDER, exist_ok=True)  # Cache system disabled

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Cache functions disabled - cache system removed for testing
# def get_file_hash(file_content):
#     """Generate hash for file caching"""
#     return hashlib.sha256(file_content).hexdigest()

# def get_cache_key(file_hash, fields_to_extract, additional_request=None, model=None):
#     """Generate cache key combining file hash, extraction parameters, and model"""
#     ...

# def get_cached_result(cache_key):
#     """Retrieve cached result if exists"""
#     ...

# def save_result_to_cache(cache_key, result):
#     """Save result to cache"""
#     ...

@app.route('/')
def index():
    """Serve the web interface"""
    return render_template('index.html')

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({"status": "ok"}), 200

@app.route('/ollama/status', methods=['GET'])
def ollama_status():
    """Return Ollama server status and installed models count"""
    ollama_url = os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434')
    try:
        resp = requests.get(f"{ollama_url}/api/tags", timeout=5)
        if resp.status_code != 200:
            return jsonify({
                "running": False,
                "error": f"HTTP {resp.status_code} from Ollama",
                "ollama_url": ollama_url
            }), 200
        data = resp.json()
        models = data.get('models', [])
        return jsonify({
            "running": True,
            "installed_models": len(models),
            "ollama_url": ollama_url
        }), 200
    except Exception as e:
        return jsonify({
            "running": False,
            "error": str(e),
            "ollama_url": ollama_url
        }), 200

@app.route('/models/available', methods=['GET'])
def get_available_models():
    """Get ONLY vision-capable models installed in Ollama"""
    ollama_url = os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434')
    try:
        response = requests.get(f"{ollama_url}/api/tags", timeout=5)
        if response.status_code != 200:
            return jsonify({"error": "Failed to fetch models from Ollama"}), 500
        
        data = response.json()
        installed_models = data.get('models', [])
        
        # Filter and return ONLY vision models
        vision_models = []
        for model in installed_models:
            model_name = model.get('name', '')
            details = model.get('details', {})
            families = details.get('families', [])
            
            # Check if vision-capable using same logic as is_vision_model()
            is_vision = False
            
            # Method 1: Check if 'clip' in families (most reliable)
            if 'clip' in families:
                is_vision = True
            
            # Method 2: Try to get detailed info from /api/show
            if not is_vision:
                try:
                    show_response = requests.post(
                        f"{ollama_url}/api/show",
                        json={"name": model_name},
                        timeout=3
                    )
                    if show_response.status_code == 200:
                        show_data = show_response.json()
                        # Check for projector_info (multimodal projector)
                        if 'projector_info' in show_data:
                            is_vision = True
                        # Double check families from show response
                        show_details = show_data.get('details', {})
                        show_families = show_details.get('families', [])
                        if 'clip' in show_families:
                            is_vision = True
                except:
                    pass  # Ignore errors, fallback to other checks
            
            # Method 3: Fallback - check known vision families
            if not is_vision:
                family = details.get('family', '').lower()
                vision_families = ['mllama', 'llava', 'bakllava', 'gemma3', 'gemma2']
                if family in vision_families or any(f in vision_families for f in families):
                    is_vision = True
            
            # Method 4: Last resort - check model name keywords
            if not is_vision:
                model_lower = model_name.lower()
                if any(keyword in model_lower for keyword in ['vision', 'llava', 'bakllava']):
                    is_vision = True
            
            # Only add vision models to the list
            if is_vision:
                vision_models.append({
                    'name': model_name,
                    'size': model.get('size', 0),
                    'modified_at': model.get('modified_at', ''),
                    'parameter_size': details.get('parameter_size', 'Unknown'),
                    'family': details.get('family', 'Unknown'),
                    'quantization': details.get('quantization_level', 'Unknown'),
                    'is_vision': True
                })
        
        return jsonify({'models': vision_models}), 200
    
    except requests.exceptions.RequestException as e:
        return jsonify({"error": f"Could not connect to Ollama: {str(e)}"}), 500

@app.route('/models/catalog', methods=['GET'])
def get_models_catalog():
    """Get the complete catalog of recommended vision models"""
    return jsonify({
        "models": VISION_MODELS,
        "hardware_recommendations": HARDWARE_RECOMMENDATIONS
    }), 200

@app.route('/models/families', methods=['GET'])
def get_models_families():
    """Return model families (Gemma3, Llama Vision, LLaVA) with small/medium/large tiers and installed flags"""
    try:
        ollama_url = os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434')
        installed_models = []
        try:
            resp = requests.get(f"{ollama_url}/api/tags", timeout=5)
            if resp.status_code == 200:
                installed_models = resp.json().get('models', [])
        except Exception:
            pass

        families = get_family_tiers()

        # Build response merging VISION_MODELS metadata where possible
        # Create a quick lookup for metadata by name
        meta_by_name = {}
        for cat in VISION_MODELS.values():
            for m in cat.get('models', []):
                meta_by_name[m['name']] = m

        def is_model_installed(target_model):
            """Check if a specific model variant is installed, using size hints when available"""
            # Extract base name and size tag from target
            target_parts = target_model.split(':')
            target_base = target_parts[0]
            target_tag = target_parts[1] if len(target_parts) > 1 else None
            
            # Size mapping: tag -> approximate parameter size (B)
            size_hints = {
                '270m': 0.3,
                '3.8b': 3.8,
                '4b': 4.0,
                '7b': 7.0,
                '11b': 11.0,
                '12b': 12.0,
                '13b': 13.0,
                '27b': 27.0,
                '90b': 90.0
            }
            
            for installed in installed_models:
                installed_name = installed.get('name', '')
                installed_parts = installed_name.split(':')
                installed_base = installed_parts[0]
                installed_tag = installed_parts[1] if len(installed_parts) > 1 else 'latest'
                
                # Must match base name
                if installed_base != target_base:
                    continue
                
                # If exact match (including tag), it's definitely installed
                if installed_name == target_model:
                    return True
                
                # If installed tag matches target tag (ignoring "latest")
                if target_tag and installed_tag == target_tag:
                    return True
                
                # IMPORTANT: Check parameter size from details for "latest" tags
                # This handles cases like llama3.2-vision:latest vs llama3.2-vision:11b
                details = installed.get('details', {})
                param_size_str = details.get('parameter_size', '')
                
                if param_size_str and target_tag:
                    # Parse parameter size like "10.7B" or "27.4B"
                    try:
                        param_size = float(param_size_str.replace('B', '').strip())
                        target_size = size_hints.get(target_tag.lower())
                        
                        if target_size:
                            # Allow ±25% tolerance for size matching (more lenient)
                            if abs(param_size - target_size) / target_size < 0.25:
                                return True
                    except (ValueError, ZeroDivisionError):
                        pass
                
                # If target has no tag and we found matching base, consider it installed
                if not target_tag and installed_base == target_base:
                    return True
            
            return False

        result = {}
        for key, fam in families.items():
            tiers = {}
            for tier, model_name in fam.get('tiers', {}).items():
                if model_name is None:
                    # Tier not available for this family
                    tiers[tier] = {
                        'model': None,
                        'installed': False,
                        'display_name': 'N/A',
                        'size': None,
                        'vram': None,
                        'speed': None,
                        'accuracy': None,
                        'description': None,
                        'use_case': None
                    }
                else:
                    meta = meta_by_name.get(model_name, {})
                    tiers[tier] = {
                        'model': model_name,
                        'installed': is_model_installed(model_name),
                        'display_name': meta.get('display_name', model_name),
                        'size': meta.get('size'),
                        'vram': meta.get('vram'),
                        'speed': meta.get('speed'),
                        'accuracy': meta.get('accuracy'),
                        'description': meta.get('description'),
                        'use_case': meta.get('use_case')
                    }
            result[key] = {
                'label': fam.get('label', key),
                'tiers': tiers
            }

        return jsonify({'families': result}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to build families: {str(e)}"}), 500

@app.route('/models/pull', methods=['POST'])
def pull_model():
    """Trigger an Ollama model pull by name. Returns aggregated progress."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        model = data.get('model')
        set_current = bool(data.get('set_current', True))
        if not model:
            return jsonify({"error": "Missing 'model' in body"}), 400

        print(f"\n{'='*60}")
        print(f"🔽 Starting download: {model}")
        print(f"{'='*60}\n")

        ollama_url = os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434')
        pull_url = f"{ollama_url}/api/pull"

        # Stream pull progress from Ollama
        progress = []
        last_status = None
        with requests.post(pull_url, json={"name": model}, stream=True, timeout=600) as r:
            r.raise_for_status()
            for line in r.iter_lines(decode_unicode=True):
                if not line:
                    continue
                # Each line is JSON like {"status":"pulling ...","completed":123,"total":456}
                try:
                    entry = json.loads(line)
                    progress.append(entry)
                    
                    # Log progress to terminal
                    status = entry.get('status', '')
                    if status != last_status:
                        print(f"📦 {status}")
                        last_status = status
                    
                    # Show progress bar for downloads
                    if 'completed' in entry and 'total' in entry:
                        completed = entry['completed']
                        total = entry['total']
                        if total > 0:
                            percent = (completed / total) * 100
                            bar_length = 40
                            filled = int(bar_length * completed / total)
                            bar = '█' * filled + '░' * (bar_length - filled)
                            print(f"\r   [{bar}] {percent:.1f}% ({completed/1024/1024:.1f}/{total/1024/1024:.1f} MB)", end='', flush=True)
                except Exception:
                    progress.append({"raw": line})

        print(f"\n\n✅ Download completed: {model}\n")

        # Optionally set current model in env for this process
        if set_current:
            os.environ['OLLAMA_MODEL'] = model
            print(f"✓ Current model set to: {model}\n")

        return jsonify({
            "success": True,
            "model": model,
            "set_current": set_current,
            "progress": progress[-10:] if len(progress) > 10 else progress  # return tail
        }), 200
    except requests.exceptions.RequestException as re:
        error_msg = f"Ollama pull failed: {str(re)}"
        print(f"\n❌ {error_msg}\n")
        return jsonify({"error": error_msg}), 500
    except Exception as e:
        error_msg = f"Unexpected error: {str(e)}"
        print(f"\n❌ {error_msg}\n")
        return jsonify({"error": error_msg}), 500

@app.route('/models/recommended', methods=['GET'])
def get_recommended_models():
    """Get recommended models based on VRAM (if provided)"""
    vram = request.args.get('vram', type=int, default=0)
    
    if vram > 0:
        recommendations = get_recommended_models_by_vram(vram)
        return jsonify(recommendations), 200
    else:
        # Return all categories
        return jsonify({
            "categories": list(VISION_MODELS.keys()),
            "hardware_recommendations": HARDWARE_RECOMMENDATIONS
        }), 200

@app.route('/models/current', methods=['GET'])
def get_current_model():
    """Get currently configured model"""
    current_model = os.environ.get('OLLAMA_MODEL', 'llama3.2-vision')
    return jsonify({
        "current_model": current_model,
        "ollama_url": os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434')
    }), 200

@app.route('/models/set', methods=['POST'])
def set_model():
    """Set the model to use for extraction"""
    data = request.get_json()
    model_name = data.get('model')
    
    if not model_name:
        return jsonify({"error": "Model name is required"}), 400
    
    # Normalize model name to match installed version
    # e.g., llama3.2-vision:11b -> llama3.2-vision:latest if that's what's installed
    normalized_model = normalize_model_name(model_name)
    
    # Update environment variable for current session
    os.environ['OLLAMA_MODEL'] = normalized_model
    
    print(f"Model set to: {normalized_model} (requested: {model_name})")
    
    return jsonify({
        "success": True,
        "model": normalized_model,
        "requested": model_name
    }), 200

def normalize_model_name(target_model):
    """
    Normalize a model name to match what's actually installed in Ollama.
    e.g., llama3.2-vision:11b -> llama3.2-vision:latest (if that's what's installed)
    """
    ollama_url = os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434')
    
    try:
        response = requests.get(f"{ollama_url}/api/tags", timeout=5)
        if response.status_code != 200:
            print(f"Warning: Could not fetch Ollama models for normalization")
            return target_model
        
        data = response.json()
        installed_models = data.get('models', [])
        
        # Parse target
        target_parts = target_model.split(':')
        target_base = target_parts[0]
        target_tag = target_parts[1] if len(target_parts) > 1 else None
        
        # Size hints for matching
        size_hints = {
            '270m': 0.3, '3.8b': 3.8, '4b': 4.0, '7b': 7.0,
            '11b': 11.0, '12b': 12.0, '13b': 13.0, '27b': 27.0, '90b': 90.0
        }
        
        # First pass: exact match
        for installed in installed_models:
            if installed.get('name') == target_model:
                return target_model
        
        # Second pass: base name match + size matching for :latest tags
        for installed in installed_models:
            installed_name = installed.get('name', '')
            installed_parts = installed_name.split(':')
            installed_base = installed_parts[0]
            
            if installed_base != target_base:
                continue
            
            # Check parameter size if target has a size tag
            if target_tag:
                details = installed.get('details', {})
                param_size_str = details.get('parameter_size', '')
                
                if param_size_str:
                    try:
                        param_size = float(param_size_str.replace('B', '').strip())
                        target_size = size_hints.get(target_tag.lower())
                        
                        if target_size and abs(param_size - target_size) / target_size < 0.25:
                            print(f"Normalized {target_model} -> {installed_name} (size match: {param_size}B ≈ {target_size}B)")
                            return installed_name
                    except (ValueError, ZeroDivisionError):
                        pass
        
        # No match found, return original
        print(f"Warning: Could not normalize {target_model}, using as-is")
        return target_model
        
    except Exception as e:
        print(f"Error normalizing model name: {e}")
        return target_model
    return jsonify({
        "success": True,
        "model": model_name,
        "message": "Model updated for current session"
    }), 200

# Cache endpoints disabled - cache system removed for testing
# @app.route('/cache/clear', methods=['POST'])
# def clear_cache():
#     """Clear all cached results"""
#     ...

# @app.route('/cache/stats', methods=['GET'])
# def cache_stats():
#     """Get cache statistics"""
#     ...

@app.route('/extract', methods=['POST'])
def extract_data_route():
    """Main extraction endpoint"""
    # Check if files are provided
    if 'file' not in request.files and 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files') if 'files' in request.files else [request.files['file']]
    if not files or all(not f.filename for f in files):
        return jsonify({"error": "No files selected"}), 400

    # Check fields_to_extract
    fields_to_extract = request.form.get('fields_to_extract')
    if not fields_to_extract:
        return jsonify({"error": "Field 'fields_to_extract' is required"}), 400
    
    try:
        fields_to_extract = json.loads(fields_to_extract)
        if not isinstance(fields_to_extract, dict):
            raise ValueError
    except Exception:
        return jsonify({"error": "Invalid format for 'fields_to_extract'. Must be JSON {field: description}"}), 400

    additional_request = request.form.get('additional_request', None)

    results = []
    try:
        for file in files:
            if file and allowed_file(file.filename):
                print(f"Processing file: {file.filename}")

                # Save file temporarily
                filename = secure_filename(file.filename)
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                file.save(filepath)

                try:
                    # Process document (extract text/image)
                    document_content = process_document(filepath, file.mimetype)
                    
                    # Extract structured data with Ollama
                    extraction_result = extract_structured_data_with_ollama(
                        document_content=document_content,
                        fields_to_extract=fields_to_extract,
                        additional_request=additional_request
                    )

                    results.append({
                        "filename": filename,
                        "extraction": extraction_result
                    })

                finally:
                    # Clean up temp file
                    if os.path.exists(filepath):
                        os.remove(filepath)
            else:
                print(f"Invalid or not allowed file skipped: {file.filename}")

        return jsonify(results), 200

    except Exception as e:
        import traceback
        print(f"Error in /extract endpoint: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/export-excel', methods=['POST'])
def export_excel():
    """Export extraction results to Excel file"""
    try:
        data = request.get_json()
        results = data.get('results', [])
        
        if not results:
            return jsonify({"error": "No results to export"}), 400
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extraction Results"
        
        # Collect all unique fields from all results
        all_fields = set()
        for result in results:
            extraction_data = result.get('extraction', {}).get('extraction_results', {}).get('data', {})
            all_fields.update(extraction_data.keys())
        
        # Sort fields for consistent column order
        sorted_fields = sorted(all_fields)
        
        # Create header row
        headers = ['Filename'] + sorted_fields + ['Confidence Score', 'Reasoning']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for result in results:
            filename = result.get('filename', 'Unknown')
            extraction = result.get('extraction', {}).get('extraction_results', {})
            data_dict = extraction.get('data', {})
            confidence = extraction.get('confidence_score', 'N/A')
            reasoning = extraction.get('reasoning', 'N/A')
            
            # Build row
            row = [filename]
            for field in sorted_fields:
                value = data_dict.get(field, '')
                # Handle None values
                if value is None:
                    value = ''
                row.append(value)
            row.append(confidence)
            row.append(reasoning)
            
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"extraction_results_{timestamp}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        print(f"Error in /export-excel endpoint: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": f"Failed to export Excel: {str(e)}"}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"Starting Flask server on port {port}...")
    app.run(debug=True, host='0.0.0.0', port=port)
